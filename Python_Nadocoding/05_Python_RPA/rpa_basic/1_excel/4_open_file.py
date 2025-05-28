from openpyxl import load_workbook  # 파일 불러오기
wb = load_workbook("sample.xlsx")   # sample.xlsx 파일에서 wb를 불러옴
ws = wb.active                      # 활성화된 sheet를 ws 변수에 할당

# cell 데이터 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=" ")  # result : 1 2 3 4 ...
#     print()

# 만약 가져올 파일의 cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):   # range에 사용하기 위해 row와 column의 max 값에 + 1을 해준다.
        print(ws.cell(row=x, column=y).value, end=" ")  # result : 1 2 3 4 ...
    print()
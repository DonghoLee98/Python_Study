from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학 -> 에서 아래처럼 '국어' column을 추가
# 번호 (국어) 영어 수학

# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "국어"     # B1 셀에 "국어" 입력

# 번호 영어 수학
ws.move_range("C1:C11", rows=5, cols=-1)    # <- 옮겨질 자리에 값이 있으면 그대로 덮어쓰기 됨에 주의

wb.save("sample_korean.xlsx")
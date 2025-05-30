from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식을 그대로 가져온다 (데이터 계산 없이 수식만)
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# load_workbook에서 data_only=True 하여 수식 계산된 후의 데이터 값만 가져온다
# evaluate 되지 않은 상태의 데이터는 None 이라고 표시
for row in ws.values:
    for cell in row:
        print(cell)

from openpyxl import Workbook
wb = Workbook()
# ws = wb.active ,,, <- 활성화된 sheet 를 ws에 담음
ws = wb.create_sheet()  # 새로운 sheet, 기본 이름으로 생성
ws.title = "MySheet"    # sheet 이름 변경

# RGB 색상 값 가져오기 : https://www.w3schools.com/colors/colors_picker.asp?colorhex=ff0000
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값 넣는다. 탭의 색상을 변경.

# Sheet, MySheet, YourSheet,,, -> Index 형태로 작성 (0, 1, 2, ...)
ws1 = wb.create_sheet("YourSheet")      # 주어진 이름으로 새로운 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)    # 2 번째 Index 자리에 해당 sheet 생성

# Sheet에 접근할 때, wb["sheet이름"] 이렇게도 가능하다
new_ws = wb["NewSheet"] # Dictionary 형태로 sheet에 접근

print(wb.sheetnames)    # Workbook 아래 모든 sheet들의 이름을 출력

# Sheet 복사
new_ws["A1"] = "Test"   # A1 셀에 "Test" 라는 값을 넣어라!
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")  # 기존 동일한 파일 명이 존재한다면, 그대로 덮어쓰기 됨.
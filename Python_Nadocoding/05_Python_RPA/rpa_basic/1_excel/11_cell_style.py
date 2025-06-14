from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"]   # 번호
b1 = ws["B1"]   # 영어
c1 = ws["C1"]   # 수학

# A column의 너비를 5로 지정
ws.column_dimensions["A"].width = 5

# 1 행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True)      # 글자 색 red, 이탤릭체, 볼드 적용
b1.font = Font(color="CC33FF", name="Arial", strike=True)   # 폰트 색상, 폰트, 취소선 적용
c1.font = Font(color="0000FF", size=20, underline="single") # 글자 크기 20, 밑줄 적용

# 셀 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 셀 배경 수정 (점수 80점 이상 학생 시나리오)
for row in ws.rows:
    for cell in row:
        # 각 cell에 대해 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # center, left, right, top, bottom 

        if cell.column == 1:    # A열(= 번호열)은 제외
            continue

        # cell의 데이터가 정수형 데이터(isinstance)이고, 그 데이터가 80보다 크면
        if isinstance(cell.value, int) and cell.value > 80:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")    # 배경은 초록색
            cell.font = Font(color="FF0000")    # 글자는 빨간색

# 틀 고정
ws.freeze_panes = "B2"  # B2 기준으로 A column, 1 row 가 scroll시 고정된다

wb.save("sample_style.xlsx")

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 셀 병합하기
ws.merge_cells("B2:D2") # B2 ~ D2 사이의 cell을 병합하겠다
ws["B2"].value = "Merged Cell"

wb.save("sample_merge.xlsx")
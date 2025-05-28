from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "DHSheet"

# A1 셀부터 해서 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])         # A1 셀의 정보를 출력   <- 값이 아닌 객체가 출력됨
print(ws["A1"].value)   # A1 셀의 값(value)을 출력
print(ws["A10"].value)  # 셀에 값이 없다면 "None" 출력

# row = 1, 2, 3, ...
# coloum = A, B, C, ... == 1, 2, 3, ... 1대1 매핑
# ws.cell(row=1, colomn=1)  # 이것은 ws["A1"]과 동일하다
print(ws.cell(column=1, row=1).value)   # = print(ws["A1"].value)
print(ws.cell(row=1, column=2).value)   # = print(ws["B1"].value)

c = ws.cell(column=3, row=1, value=10)  # ws["C1"].value = 10 과 동일
print(c.value)  # ws["C1"]

from random import *

# 반복문을 이용하여 랜덤 숫자 채우기
index = 1
for x in range(1, 11):  # 10 개의 row
    for y in range(1, 11):  # 10 개의 column
        # ws.cell(row=x, column=y, value=randint(0, 100)) # 0 ~ 100 사이의 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample.xlsx")
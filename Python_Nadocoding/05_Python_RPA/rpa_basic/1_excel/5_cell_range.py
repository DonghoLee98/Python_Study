from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 한 줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):  # 10 개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 영어 column(= B column)만 가지고 오기
# # print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"]   # 영어, 수학 column 함께 가지고 오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1]   # 1번째 row만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # 2 ~ 6번째 row의 값 가지고 오기 (1번째 row인 title 제외)
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()   # 줄바꿈용!

from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row]    # 2번째 부터 마지막 row 까지의 데이터
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")        # cell의 value 가져옴
#         # print(cell.coordinate, end=" ")     # cell의 coordinate 가져옴
#         xy = coordinate_from_string(cell.coordinate)    # tuple 형태로, row와 column을 분리해서 return 한다
#         # print(xy, end=" ")  # row : int / column : string
#         print(xy[0], end="")   # A
#         print(xy[1], end=" ")   # 1
#     print()

# 전체 rows
# print(tuple(ws.rows))       # 한 row가 tuple의 기준
# for row in tuple(ws.rows):
#     print(row[1].value)

# 전체 columns
# print(tuple(ws.columns))    # 한 column이 tuple의 기준
# for col in tuple(ws.columns):
#     print(col[0].value)

# for row in ws.iter_rows():  # 전체 row에 대해 반복
#     print(row[1].value)

# for col in ws.iter_cols():
#     print(col[2].value)

# 2 ~ 11번째 row와 2 ~ 3번째 column의 데이터를 끊어서 가져온다
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):     # iter_rows : 좌에서 우로 데이터 가져옴
    # print(row[0].value, row[1].value)   # [0]수학, [1]영어
    # print(row)

# 범위 지정할 때 min, max 값 따로 명시하지 않으면 최소, 최댓값을 알아서 가져온다
for col in ws.iter_cols(min_row=2):        # iter_cols : 위에서 아래로 데이터 가져옴
    print(col.val)

wb.save("sample.xlsx")